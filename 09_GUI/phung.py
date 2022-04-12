import string
import sys
from PyQt5 import  QtWidgets
from PyQt5 import uic
import serial 
import serial.tools.list_ports
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from PyQt5 import QtCore
from PyQt5.QtCore import pyqtSlot
from Draw import *
import time
import openpyxl
import numpy as np

# Phân Luồng xử lý
class Worker(QtCore.QRunnable):

	def __init__(self, function, *args, **kwargs):
		super(Worker, self).__init__()
		self.function = function
		self.args = args
		self.kwargs = kwargs

	@pyqtSlot()
	def run(self): 
		self.function(*self.args, **self.kwargs)           

# Đối tượng chính của giao diện
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui= uic.loadUi("mainwindow.ui",self)
        self.ser =  serial.Serial()
        self.theta=[]
        self.theta_1=[]
        self.theta_2=[]
        self.theta_3=[]
        self.theta_4=[]
        self.time_dem=[]
        self.ennable_receive=0
        self.time_dem_tam=0
        self.line_data=1
        self.threadpool = QtCore.QThreadPool() 
        self.list_port()                      # Chạy hàm này để cập nhật port được kết nối

        self.ui.actionHome.triggered.connect(self.showscreen1)
        self.ui.actionParameter.triggered.connect(self.showscreen2)
        self.ui.actionController.triggered.connect(self.showscreen3)
        self.ui.actionPositions.triggered.connect(self.showscreen4)
        self.ui.actionErrors.triggered.connect(self.showscreen5)
        self.ui.btn_connect_stm.clicked.connect(self.connect_stm_clicked)
        self.ui.btn_disconnect_stm.clicked.connect(self.disconnect_stm_clicked)
        self.ui.btn_connect_start_receive.clicked.connect(self.start_receive)
        self.ui.btn_connect_stop_receive.clicked.connect(self.stop_receive)
        # Nhúng đồ thị vào màn hình position
        self.screen_theta1 = Display_theta1(width=100, height=40, dpi=70)
        self.screen_theta1.config_display_theta1(self.screen_theta1)
        self.ui.plot_the1.addWidget(self.screen_theta1)

        self.screen_theta2 = Display_theta2(width=100, height=40, dpi=70)
        self.screen_theta2.config_display_theta2(self.screen_theta2)
        self.ui.plot_the2.addWidget(self.screen_theta2)

        self.screen_theta3 = Display_theta3(width=100, height=40, dpi=70)
        self.screen_theta3.config_display_theta3(self.screen_theta3)
        self.ui.plot_the3.addWidget(self.screen_theta3)

        self.screen_theta4 = Display_theta4(width=100, height=40, dpi=70)
        self.screen_theta4.config_display_theta4(self.screen_theta4)
        self.ui.plot_the4.addWidget(self.screen_theta4)

        # Nhúng đồ thị vào màn hình error
        self.screen_error1 = Display_error1(width=100, height=40, dpi=70)
        self.screen_error1.config_display_error1(self.screen_error1)
        self.ui.plot_error1.addWidget(self.screen_error1)

        self.screen_error2 = Display_error2(width=100, height=40, dpi=70)
        self.screen_error2.config_display_error2(self.screen_error2)
        self.ui.plot_error2.addWidget(self.screen_error2)

        self.screen_error3 = Display_error3(width=100, height=40, dpi=70)
        self.screen_error3.config_display_error3(self.screen_error3)
        self.ui.plot_error3.addWidget(self.screen_error3)

        self.screen_error4 = Display_error4(width=100, height=40, dpi=70)
        self.screen_error4.config_display_error4(self.screen_error4)
        self.ui.plot_error4.addWidget(self.screen_error4)
        
        # vẽ và lưu dữ liệu vào excel
        self.start_draw()

    # Hàm vẽ đồ thị góc
    def draw_theta(self): 
        while True:
            self.screen_theta1.axes.clear()
            self.screen_theta1.config_display_theta1(self.screen_theta1)    
            self.screen_theta1.axes.plot(self.time_dem, self.theta_1, linewidth=2)
            self.screen_theta1.draw()
            self.screen_theta2.axes.clear()
            self.screen_theta2.config_display_theta2(self.screen_theta2)    
            self.screen_theta2.axes.plot(self.time_dem, self.theta_2, linewidth=2)
            self.screen_theta2.draw()
            self.screen_theta3.axes.clear()
            self.screen_theta3.config_display_theta3(self.screen_theta3)    
            self.screen_theta3.axes.plot(self.time_dem, self.theta_3, linewidth=2)
            self.screen_theta3.draw()
            self.screen_theta4.axes.clear()
            self.screen_theta4.config_display_theta4(self.screen_theta4)    
            self.screen_theta4.axes.plot(self.time_dem, self.theta_4, linewidth=2)
            self.screen_theta4.draw()

    # Hàm bắt đầu luồng vẽ đồ thị, dược khởi động bằng nút nhấn
    def start_draw(self):
        worker = Worker(lambda: self.draw_theta())
        self.threadpool.start(worker)

    # Hàm ghi tên của các cột vào file excel
    def write_data_excel(self):
        wb=openpyxl.load_workbook('Data.xlsx')
        sheet1= wb['Sheet1']
        sheet1["A1"].value="the_1"
        sheet1["B1"].value="the_1"
        sheet1["C1"].value="the_3"
        sheet1["D1"].value="the_4"
        wb.close
        wb.save('Data.xlsx') 

    # Hàm nhận giữ liệu từ serial
    def receive(self):
        while self.ser.isOpen() and self.ennable_receive==1:
            time.sleep(0.1)
            data = self.ser.readline().decode()
            self.ser.flushInput()
            self.theta = data.strip().split(",")
            if len(self.theta) >= 7 and self.theta[0]=='in' and self.theta[6]=='fi':
                self.theta_1.append(float(self.theta[1]))
                self.theta_2.append(float(self.theta[2]))
                self.theta_3.append(float(self.theta[3]))
                self.theta_4.append(float(self.theta[4]))
                self.time_dem.append(float(self.theta[5]))
                self.line_data=self.line_data+1                         # Tăng số thứ tự hàng trong excel
                
                # Ghi dữ liệu vào file excel
                cellname_1=("A"+str(self.line_data))
                cellname_2=("B"+str(self.line_data))
                cellname_3=("C"+str(self.line_data))
                cellname_4=("D"+str(self.line_data))
                wb=openpyxl.load_workbook('Data.xlsx')
                sheet1= wb['Sheet1']
                sheet1[cellname_1].value=self.theta[1]
                sheet1[cellname_2].value=self.theta[2]
                sheet1[cellname_3].value=self.theta[3]
                sheet1[cellname_4].value=self.theta[4]
                wb.close
                wb.save('Data.xlsx')

    # Hàm  bắt đầu luồng nhận giữ liệu, được khởi động bằng nút nhấn         
    def start_receive(self):
        self.ennable_receive=1
        wb=openpyxl.load_workbook('Data.xlsx')
        sheet1= wb['Sheet1']
        sheet1.delete_cols(idx=1, amount=4)
        wb.close
        wb.save('Data.xlsx')
        self.write_data_excel()
        if self.ser.isOpen() :
            self.ser.write('1'.encode())
        self.theta=[]
        self.time_dem=[]
        self.theta_1=[]
        self.theta_2=[]
        self.theta_3=[]
        self.theta_4=[]
        worker_receive = Worker(lambda: self.receive())
        self.threadpool.start(worker_receive)
        self.btn_connect_start_receive.setStyleSheet('QPushButton {background-color:#008000;font: 16pt "Times New Roman"; color: white;}') 
        self.btn_connect_stop_receive.setStyleSheet('QPushButton {background-color:#474747;font: 16pt "Times New Roman"; color: white;}')
    # Hàm yêu cầu dừng nhận giữ liệu từ serial
    def stop_receive(self):
        if self.ser.isOpen() and self.ennable_receive==1:
            self.ser.write('0'.encode())
        self.ennable_receive=0
        self.line_data=0
        self.btn_connect_stop_receive.setStyleSheet('QPushButton {background-color:#B90000;font: 16pt "Times New Roman"; color: white;}') 
        self.btn_connect_start_receive.setStyleSheet('QPushButton {background-color:#474747;font: 16pt "Times New Roman"; color: white;}')

    # Hàm hiển thị các màn hình chính của giao diện
    def showscreen1(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.screen_1)
    def showscreen2(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.screen_2)
    def showscreen3(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.screen_3)
    def showscreen4(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.screen_4)        
    def showscreen5(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.screen_5)

    # Kết nối giao diện với serial
    def list_port(self):
        ports = serial.tools.list_ports.comports()
        self.commPort =([comport.device for comport in ports])
        self.numConnection = len(self.commPort)
        if  self.numConnection == 0 :
            pass
        elif self.numConnection == 1 :
             self.port_stm.addItem(str(self.commPort[0]))
        else:
             self.port_stm.addItem(str(self.commPort[0]))
             self.port_stm.addItem(str(self.commPort[1]))

    def connect_stm_clicked(self):
        comport = self.port_stm.currentText()
        baurate = self.baud_stm.currentText()     
        self.ser.port = comport
        self.ser.baudrate =  baurate
        self.ser.bytesize = serial.EIGHTBITS #number of bits per bytes
        self.ser.parity = serial.PARITY_NONE #set parity check: no parity
        self.ser.stopbits = serial.STOPBITS_ONE #number of stop bits            #timeout block read
        self.ser.xonxoff = False     #disable software flow control
        self.ser.rtscts = False     #disable hardware (RTS/CTS) flow control
        self.ser.dsrdtr = False       #disable hardware (DSR/DTR) flow control
        self.ser.writeTimeout = 0    #timeout for write
        self.ser.timeout =0
        self.ser.open()
        if self.ser.isOpen():
            self.label_stm.setText('Connected') 
            self.btn_connect_stm.setStyleSheet('QPushButton {background-color:#008000;font: 16pt "Times New Roman"; color: white;}') 
            self.btn_disconnect_stm.setStyleSheet('QPushButton {background-color:#474747;font: 16pt "Times New Roman"; color: white;}')          
                       
    def disconnect_stm_clicked(self):
        self.ser.close()
        if not self.ser.isOpen():
            self .label_stm.setText('Disconnected')
            self.btn_connect_stm.setStyleSheet('QPushButton {background-color:#474747;font: 16pt "Times New Roman"; color: white;}') 
            self.btn_disconnect_stm.setStyleSheet('QPushButton {background-color:#B90000;font: 16pt "Times New Roman"; color: white;}')

if __name__ == '__main__':
     app= QtWidgets.QApplication(sys.argv)
     interface=MainWindow()
     interface.show()
     sys.exit(app.exec_())



        
